import csv
import openpyxl
import pprint
import glob
import os
from tabula import read_pdf
import tabula
import time
from pyexcel.cookbook import merge_all_to_a_book
import pandas as pd
import pyautogui
import collections
from selenium import webdriver
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC

# depending on your program, pyautogui can "go rogue". as failsafe,
#we may move the mouse to the upper left corner of the screen to 
#halt
pyautogui.FAILSAFE = True

def translatepdf():
    """
    use tabula to convert pdf file into .csv
    """
    list_of_files = glob.glob('/Users/shaneshimizu/Downloads/*') # * means all if need specific format then *.csv
    latest_file = max(list_of_files, key=os.path.getctime)
    validatefile = input("name this file: ")
    #validatefile = input("Is " + latest_file + " the file you want to process?" + "\n type yes or no: ")

    if validatefile.lower() != None:
        dateRange = validatefile
        #dateRange = input("Please enter the date range for this report (use underscore instead of spaces): ")
        data = read_pdf(latest_file, pages = 'all')
        tabula.convert_into(latest_file, f'Fuel_{dateRange}_EXP.csv', output_format="csv", pages = 'all')
        time.sleep(3)
        createTemplate(dateRange)
    else:
        exit()

def createTemplate(dateRange):
    """
    Create a template
    """
    #EXP
    wb = openpyxl.Workbook() 
    sheet = wb.active
    sheet.title='Sheet 1' #change title

    wb.save(f'Template_Fuel_Charges_{dateRange}_EXP.csv')
    createdServiceBilling = f'Fuel_{dateRange}_EXP.csv'
    print("EXP Template Created")

    #REV
    wb2 = openpyxl.Workbook() 
    sheet = wb2.active
    sheet.title='Sheet 1' #change title

    wb2.save(f'Template_Fuel_Charges_{dateRange}_REV.csv')
    createdServiceBilling2 = f'Fuel_{dateRange}_REV.csv'
    print("REV Template Created")
    #time.sleep(2)

    list_of_files2 = glob.glob(createdServiceBilling) # * means all if need specific format then *.csv
    latest_file2 = max(list_of_files2, key=os.path.getctime)
    os.system(f"open {latest_file2}")
    
    time.sleep(2)
    pyautogui.FAILSAFE = True
    #file (120,13)
    pyautogui.click(116,11)
    #save as (174,165)
    pyautogui.click(161,161)
    #file format (835,208)
    pyautogui.click(1808,447)
    #move to position, wait (maybe),  then click on xlsx format (850,19)
    pyautogui.moveTo(1762,119, duration = 0.25)
    time.sleep(1)
    pyautogui.click()
    #save (1030,295)
    time.sleep(2)
    pyautogui.click(1978,539)
    #close (89,51)
    time.sleep(1)
    pyautogui.click(1038,291)

    time.sleep(5)
    print("Now Executing Automation...")

    executeAutomation(createdServiceBilling)

def executeAutomation(createdServiceBilling):

    time.sleep(1)

    list_of_files3 = glob.glob('/Users/shaneshimizu/Desktop/VSCode/automationtesting/*') # * means all if need specific format then *.csv
    latest_file3 = max(list_of_files3, key=os.path.getctime)
    #print(latest_file3)
    print("Calculating Totals...")
    #OPEN WORKBOOK
    try:
        fname = latest_file3
        wb = openpyxl.load_workbook(fname)
    except:
        print("need to save file as xlsx")
        return
    createdServiceBilling = createdServiceBilling[:-4]
    sheet = wb[f"{createdServiceBilling}"]

    #GRAB ALL THE CELLS WITH MA (not gas)
    listofcoordinate = []
    listofaccounts = []
    testaccountlist = []
    for rowOfCellObjects in sheet['D1':'D209']:
        for cellObj in rowOfCellObjects:
                regulartotal = "MA"
                whyaccounthere = None
                strCellObj = str(cellObj.value)
                whatsthis = cellObj.coordinate[1:4]
                bcoordinate = sheet[f"B{whatsthis}"]
                ccoordinate = sheet[f"C{whatsthis}"]
                ccoordinateFinal = str(ccoordinate.value)
                bcoordinateFinal = str(bcoordinate.value)
                regulargastotal = "MA"
                notmail = "MAIL"
                if strCellObj == "None" and regulargastotal in bcoordinateFinal:
                    MACoordinate2 = str(cellObj.coordinate)
                    MACoordinateFinal2 = MACoordinate2[1:4]
                    bcellvalue = sheet[f"B{MACoordinateFinal2}"]
                    bcelvaluefinal = bcellvalue.value[-10:]
                    sheet[f"D{MACoordinateFinal2}"] = bcelvaluefinal
                    updatedDcoordinate = sheet[f"D{MACoordinateFinal2}"]
                    listofcoordinate.append(updatedDcoordinate.coordinate)
                    listofaccounts.append(updatedDcoordinate.coordinate)
                    wb.save(fname)
                if strCellObj == "0" and regulargastotal in ccoordinateFinal:
                    MACoordinateC = str(cellObj.coordinate)
                    MACoordinateFinalC = MACoordinateC[1:4]
                    ccellvalue = sheet[f"C{MACoordinateFinalC}"]
                    ccellvaluefinal = ccellvalue.value
                    sheet[f"D{MACoordinateFinalC}"] = ccellvaluefinal[-10:]
                    updatedCcoordinate = sheet[f"D{MACoordinateFinalC}"]
                    listofcoordinate.append(updatedCcoordinate.coordinate)
                    listofaccounts.append(updatedCcoordinate.coordinate)
                    wb.save(fname)
                elif regulargastotal in ccoordinateFinal and notmail not in ccoordinateFinal:
                    MACoordinateC = str(cellObj.coordinate)
                    #should grab end of string
                    MACoordinateFinalC = MACoordinateC[1:4]
                    ccellvalue = sheet[f"C{MACoordinateFinalC}"]
                    ccellvaluefinal = ccellvalue.value
                    sheet[f"D{MACoordinateFinalC}"] = ccellvaluefinal[-10:]
                    updatedCcoordinate = sheet[f"D{MACoordinateFinalC}"]
                    listofcoordinate.append(updatedCcoordinate.coordinate)
                    listofaccounts.append(updatedCcoordinate.coordinate)
                    wb.save(fname)
                if regulartotal in strCellObj:
                    MACoordinate = cellObj.coordinate
                    MAValue = cellObj.value
                    listofcoordinate.append(MACoordinate)
                    listofaccounts.append(MACoordinate)
                    wb.save(fname)

    #GRAB ALL THE CELLS WITH MA (only gas) and replace B w/ D
    listofGascoordinates = []
    listofgasaccounts = []
    for rowOfCellObjects in sheet['B1':'B209']:
        for cellObj in rowOfCellObjects:
            regulartotal = "GAS"
            strCellObj = str(cellObj.value)
            if regulartotal in strCellObj:
                MACoordinate = str(cellObj.coordinate)
                MACoordinateFinal = MACoordinate[1:4]
                MAValue = cellObj.value
                MAAccount = cellObj.value
                listofGascoordinates.append(f"D{MACoordinateFinal}")
                listofaccounts.remove(f"D{MACoordinateFinal}")

    #get account D if B has gas    
    for accountCells in sheet[f'D1': 'D209']:
        for cellobjfinal in accountCells:
            strcellobjfinal = cellobjfinal.coordinate
            strcellobjfinalval = cellobjfinal.value
            y = 0
            while y < len(listofGascoordinates):
                if listofGascoordinates[y] == strcellobjfinal:
                    listofgasaccounts.append(f"{strcellobjfinalval}G")
                y += 1
    
    #get prices for the Gas only
    h = 0
    listofgasprices = []
    listOfFuelPrices = []
    while h < len(listofGascoordinates):
        getValOnly2 = listofGascoordinates[h]
        finalVal2 = getValOnly2[1:4]
        thecells8 = sheet[f'K{finalVal2}']
        if thecells8.value == None:
            print("no vlaue here")
            #set K cell equal to J cell
            sheet[f"K{finalVal2}"].value = sheet[f"J{finalVal2}"].value
            listofgasprices.append(sheet[f"K{finalVal2}"].value)
            wb.save(fname)
        else:
            fuelpriceCoor8 = thecells8.coordinate
            fuelpriceVal8 = thecells8.value
            listofgasprices.append(fuelpriceVal8)
        h += 1

    #REMOVES ALL account COORDINATES that have the gas
    matchingvalues = []
    c = 0
    while c < len(listofGascoordinates):
        if any(thiscoor in listofcoordinate for thiscoor in listofGascoordinates):
            listofcoordinate.remove(str(listofGascoordinates[c]))
        c += 1


    #GET THE FUEL PRICES after removing gas coordinates
    i = 0
    #listOfFuelPrices = []
    while i < len(listofcoordinate):
        getValOnly = listofcoordinate[i]
        finalVal = getValOnly[1:4]
        thecells = sheet[f'K{finalVal}']
        if thecells.value == None:
            print("no vlaue here")
            #set K cell equal to J cell
            sheet[f"K{finalVal}"].value = sheet[f"J{finalVal}"].value
            listOfFuelPrices.append(sheet[f"K{finalVal}"].value)
            wb.save(fname)
        else:
            fuelpriceCoor = thecells.coordinate
            fuelpriceVal = thecells.value
            listOfFuelPrices.append(fuelpriceVal)
        i += 1


    #GET ACCOUNT NUMBER AFTER REMOVING gas
    r = 0
    actualaccountlist = []
    while r < len(listofcoordinate):
        getValOnly2 = listofcoordinate[r]
        finalVal2 = getValOnly2[1:4]
        thecells2 = sheet[f'D{finalVal2}']
        accountcoor2 = thecells2.coordinate
        accountVal2 = str(thecells2.value)
        actualaccountlist.append(accountVal2)
        r += 1


    #FIND EACH UNIQUE INSTANCE OF REGULAR FUEL and put into list
    dicEachAccountTotals = {}
    uniqueAccounts = set(actualaccountlist)
    uniqueAccountsList = list(uniqueAccounts)
    p = 0
    while p < len(uniqueAccountsList):
        dicEachAccountTotals[f"{uniqueAccountsList[p]}"] = 0
        p += 1

    #FIND EACH UNIQUE INSTANCE OF GAS and put into list
    dicEachGasAccountTotals = {}
    uniqueGasAccounts = set(listofgasaccounts)
    uniqueAccountsGasList = list(uniqueGasAccounts)
    s = 0
    while s < len(uniqueAccountsGasList):
        dicEachGasAccountTotals[f"{uniqueAccountsGasList[s]}"] = 0
        s += 1


    #UPDATE THE values IN EACH ACCOUNT after adding all the prices for each account (fuel only)
    q = 0
    while q < len(actualaccountlist):
        if actualaccountlist[q] in dicEachAccountTotals:
            dicEachAccountTotals[f"{actualaccountlist[q]}"] = []
        q += 1

    #UPDATE THE values IN EACH ACCOUNT after adding all the GAS PRICES for each account (gas only)
    u = 0
    while u < len(listofgasaccounts):
        if listofgasaccounts[u] in dicEachGasAccountTotals:
            dicEachGasAccountTotals[f"{listofgasaccounts[u]}"] = []
        u += 1

    #MERGE ACCOUNT AND FUEL LIST
    combinedFuelList = list(zip(actualaccountlist, listOfFuelPrices))
    combinedGasList = list(zip(listofgasaccounts, listofgasprices))

    #PRINT THE 1ST/2ND INDEX OF TUPLE IN LIST
    firstTuple = [u[0] for u in combinedFuelList]
    secondTuple = [u[1] for u in combinedFuelList]

    firstTupleGas = [w[0] for w in combinedGasList]
    secondTupleGas = [w[1] for w in combinedGasList]


    
    #APPEND RESPECTIVE CHARGE TO EACH ACCOUNT (fuel only)
    y = 0
    while y < len(combinedFuelList):
        if firstTuple[y] in dicEachAccountTotals.keys():
            dicEachAccountTotals[firstTuple[y]].append(secondTuple[y])
        y += 1

    #APPEND RESPECTIVE CHARGE TO EACH ACCOUNT (gas only)
    w = 0
    while w < len(combinedGasList):
        if firstTupleGas[w] in dicEachGasAccountTotals.keys():
            dicEachGasAccountTotals[firstTupleGas[w]].append(secondTupleGas[w])
        w += 1
    #GET THE SUM OF EACH ACCOUNT (fuel only)
    pprint.pprint(dicEachAccountTotals)
    for key2 in dicEachAccountTotals:
        try:
            dicEachAccountTotals[key2] = sum(dicEachAccountTotals[key2])
        except:
            print("empty cell in fuel prices")
            return
    for key3 in dicEachGasAccountTotals:
        dicEachGasAccountTotals[key3] = sum(dicEachGasAccountTotals[key3])

    dicEachAccountTotals.update(dicEachGasAccountTotals)
    finaldicEachAccountTotals = collections.OrderedDict(sorted(dicEachAccountTotals.items()))
    #pprint.pprint(dicEachAccountTotals)

    sendToExcel(finaldicEachAccountTotals)


def sendToExcel(finaldicEachAccountTotals):
    list_of_files5 = glob.glob('/Users/shaneshimizu/Desktop/VSCode/automationtesting/*') # * means all if need specific format then *.csv
    latest_file5 = sorted(list_of_files5, key=os.path.getctime)
    print("EXP File saved as:", latest_file5[-4])
    print("REV File saved as:", latest_file5[-3])
    filenameEXP = latest_file5[-4]
    filenameREV = latest_file5[-3]
    with open(filenameEXP, "w") as f:
        writer = csv.writer(f, dialect="excel")
        hasWHSE = False
        for key, value in finaldicEachAccountTotals.items():
            if key[-1] == 'G':
                gasCode = 3035
                writer.writerow(('MA', key[3:10], '', gasCode, '', '', '', 'GAS CHARGE', round(value, 2)))
            else:
                gasCode = 3025
                writer.writerow(('MA', key[3:10], '', gasCode, '', '', '', 'GAS CHARGE', round(value, 2)))
            if key[-4] == "WHSE":
                hasWHSE = True
    
    with open(filenameREV, "w") as g:
        writer = csv.writer(g, dialect="excel")
        for key, value in finaldicEachAccountTotals.items():
            writer.writerow(('MA','2221362', '', '0793', '', '', '', 'GAS CHARGE', round(value, 2)))
    
    return
    uploadKFS(filenameREV, filenameEXP, hasWHSE)
    
def uploadKFS(filenameREV, filenameEXP):
    driver = webdriver.Chrome()


    driver.get("https://kfs.hawaii.edu/kfs-prd6/portal.jsp")
    WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="iframe_portlet_container_table"]/tbody/tr[2]/td[2]/div[2]/div[2]/div/ul[5]/li[6]/a'))).click()

    WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div/div[1]/div[2]/form/div[1]/span/input'))).click()
    WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div/div[1]/div[2]/form/div[1]/span/input'))).send_keys('foo')
    WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div/div[1]/div[2]/form/div[2]/input'))).click()
    WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div/div[1]/div[2]/form/div[2]/input'))).send_keys('foo')
    WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div/div[1]/div[2]/form/p[2]/input[3]'))).click()

    WebDriverWait(driver, 20).until(EC.frame_to_be_available_and_switch_to_it((By.TAG_NAME,'iframe')))

    WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="login-form"]/div[2]/div/label/input'))).click()
    WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="auth_methods"]/fieldset/div[1]/button'))).click()
    driver.switch_to.default_content()

    time.sleep(5)
    driver.switch_to.default_content()

    WebDriverWait(driver, 20).until(EC.frame_to_be_available_and_switch_to_it((By.XPATH,'/html/body/div[8]/div/iframe')))
    iframethis2 = driver.find_element_by_id('iframeportlet')
    driver.switch_to.frame(iframethis2)

    WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div/form/table/tbody/tr/td[2]/div/div[1]/div[2]/table[1]/tbody/tr[1]/td[2]/textarea'))).click()
    WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div/form/table/tbody/tr/td[2]/div/div[1]/div[2]/table[1]/tbody/tr[1]/td[2]/textarea'))).send_keys('Fleet Services - Gas Charges for University vehicles from (September 6-12 2020)')
     #REV
    WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div/form/table/tbody/tr/td[2]/div/div[2]/div[2]/table/tbody/tr[2]/td[2]/a/img'))).click()
    WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div/form/table/tbody/tr/td[2]/div/div[2]/div[2]/table/tbody/tr[2]/td[2]/div/input[1]'))).send_keys(f"{filenameREV}")
    WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div/form/table/tbody/tr/td[2]/div/div[2]/div[2]/table/tbody/tr[2]/td[2]/div/input[2]'))).click()
    #EXP NOT WORKING
    time.sleep(5)
    print("now trying exp")
    WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div/form/table/tbody/tr/td[2]/div/div[2]/div[2]/table/tbody/tr[69]/td[2]/a/img'))).click()
    WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div/form/table/tbody/tr/td[2]/div/div[2]/div[2]/table/tbody/tr[69]/td[2]/div/input[1]'))).send_keys(f"{filenameEXP}")
    WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div/form/table/tbody/tr/td[2]/div/div[2]/div[2]/table/tbody/tr[69]/td[2]/div/input[2]'))).click()
    #if WHSE = True, input WHSE into object code




    time.sleep(10000)

    #driver.switch_to.default_content()
    

if __name__ == "__main__":
    translatepdf()
